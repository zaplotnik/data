#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parses Slovenia's 2026 CRT submission (emissions/sources/SVN_2026/) and
(re)generates emissions/data/*.csv for years 1990-2024.

Run with cwd = emissions/src/ (paths below are relative to that), in an
environment with pandas, numpy and openpyxl installed.

Background: the 2026 CRT workbooks use a different sheet layout than every
prior submission (SVN_2021.../SVN_2022/SVN_20230315) that
emissions/sources/raw2emissions.py was written against - sheet count, sheet
positions, and even the row offset within a sheet (e.g. the "Total CO2
equivalent emissions without LULUCF" row) all shifted, and not by a constant
amount. Rather than hardcoding row/column positions, this script locates each
of the ~60 known Summary2 categories (and the transport sub-categories in
Table1.A(a)s3) by matching their label text, in the fixed order those
categories are known to appear, skipping blank rows and raising immediately
if an expected category can't be found - so a future format change fails
loudly instead of silently misaligning data.

Per user decision: the data series now starts at 1990 (SVN_2026 has no data
for 1987-1989, so those - and 1986 - are dropped rather than kept from the
old submission). Overlapping years (1990-2021) are replaced with the revised
SVN_2026 figures. road_transporation.heavy_duty_trucks/.buses are derived,
for 1990-2021, from the manually-maintained split spreadsheet ("Emisije TGP
iz cestnega prometa...xlsx", same source emissions_historical_transport_xlsx2csv.py
used) and left at 0.0 for 2022-2024, since that spreadsheet's columns stop
at 2021 and it isn't part of the SVN_2026 submission. The CRT workbooks
themselves never split heavy duty trucks from buses - they report a single
"heavy duty trucks and buses" row - so that combined figure is also kept as
its own column, road_transporation.heavy_duty_trucks_and_buses, sourced
directly from each year's workbook and available for the full 1990-2024
range (unlike the split, which still stops at 2021).
"""

import glob
import os
import re
import shutil
from datetime import datetime

import numpy as np
import openpyxl as xl
import pandas as pd

YEAR_START = 1990
YEAR_END = 2024

GWP = {"co2": 1, "ch4": 25, "n2o": 298}

CODE_RE = re.compile(r"^(NA|NO|NE|IE|C)(\s*,\s*(NA|NO|NE|IE|C))*$", re.IGNORECASE)


def normalize_label(value):
    if not isinstance(value, str):
        return None
    text = re.sub(r"\s+", " ", value.strip())
    if not text or CODE_RE.match(text):
        return None
    return text.lower()


def to_number(value):
    if isinstance(value, str) or value is None:
        return 0.0
    return float(value)


def co2equiv(co2, ch4, n2o):
    return to_number(co2) * GWP["co2"] + to_number(ch4) * GWP["ch4"] + to_number(n2o) * GWP["n2o"]


def match_sequence(rows, sequence, label_cols, value_fn, path):
    """Walk `rows` (tuples of openpyxl Cells) matching each (key, keyword) in
    `sequence` against the first label found in `label_cols` of each row, in
    order. Rows whose label doesn't contain the current expected keyword are
    skipped (blank spacer rows, but also section headers like "Memo items:"
    that carry text but aren't one of the tracked categories) - categories
    must still show up in order, just not on consecutive rows. `value_fn(row)`
    extracts whatever should be stored for a matched row."""
    values = {}
    seq_iter = iter(sequence)
    key, keyword = next(seq_iter)
    for row in rows:
        label = None
        for col in label_cols:
            label = normalize_label(row[col].value)
            if label:
                break
        if label is None or keyword not in label:
            continue
        values[key] = value_fn(row)
        try:
            key, keyword = next(seq_iter)
        except StopIteration:
            return values
    raise ValueError(
        f"{path}: only matched {len(values)}/{len(sequence)} categories "
        f"(next expected: '{keyword}')"
    )


# Ordered (key, keyword) pairs for the 58 real Summary2 rows (indices 48/49 of
# the historical `ei` mapping are blank spacer rows with no label, and are
# naturally skipped since we advance only on a real, matching label).
SUMMARY2_SEQUENCE = [
    ("total_net", "total (net emissions)"),
    ("energy.total", "energy"),
    ("energy.fuel_combustion_activities.total", "fuel combustion"),
    ("energy.fuel_combustion_activities.energy_industries", "energy industries"),
    ("energy.fuel_combustion_activities.manufacturing_construction", "manufacturing industries and construction"),
    ("energy.fuel_combustion_activities.transport", "transport"),
    ("energy.fuel_combustion_activities.other_sectors", "other sectors"),
    ("energy.fuel_combustion_activities.other", "other"),
    ("energy.fugitive_emissions_from_fuels.total", "fugitive emissions from fuels"),
    ("energy.fugitive_emissions_from_fuels.solid_fuels", "solid fuels"),
    ("energy.fugitive_emissions_from_fuels.oil_natural_gas_and_energy_production", "oil and natural gas"),
    ("energy.co2_transport_storage", "co2 transport and storage"),
    ("industrial_processes.total", "industrial processes and product use"),
    ("industrial_processes.mineral_industry", "mineral industry"),
    ("industrial_processes.chemical_industry", "chemical industry"),
    ("industrial_processes.metal_industry", "metal industry"),
    ("industrial_processes.non_energy_products_from_fuels", "non-energy products from fuels"),
    ("industrial_processes.electronic_industry", "electronic industry"),
    ("industrial_processes.product_usese_as_ODS", "product uses as ods"),
    ("industrial_processes.other_product_manufacture_use", "other product manufacture and use"),
    ("industrial_processes.other", "other"),
    ("agriculture.total", "agriculture"),
    ("agriculture.enteric_fermentation", "enteric fermentation"),
    ("agriculture.manure_management", "manure management"),
    ("agriculture.rice_cultivation", "rice cultivation"),
    ("agriculture.agricultural_soils", "agricultural soils"),
    ("agriculture.prescribed_burning_of_savannas", "prescribed burning of savanna"),
    ("agriculture.field_burning_agricultural_residues", "field burning of agricultural residues"),
    ("agriculture.liming", "liming"),
    ("agriculture.urea_application", "urea application"),
    ("agriculture.carbon_containing_fertilizers", "carbon-containing fertilizers"),
    ("agriculture.other", "other"),
    ("lulucf.total", "land use, land-use change and forestry"),
    ("lulucf.forest_land", "forest land"),
    ("lulucf.cropland", "cropland"),
    ("lulucf.grassland", "grassland"),
    ("lulucf.wetlands", "wetlands"),
    ("lulucf.settlements", "settlements"),
    ("lulucf.other_land", "other land"),
    ("lulucf.harvested_wood_prducts", "harvested wood products"),
    ("lulucf.other", "other"),
    ("waste.total", "waste"),
    ("waste.solid_waste_disposal", "solid waste disposal"),
    ("waste.biological_treatment_solid_waste", "biological treatment of solid waste"),
    ("waste.incineration_open_burning_waste", "incineration and open burning of waste"),
    ("waste.waste_water_treatment_discharge", "waste water treatment and discharge"),
    ("waste.other", "other"),
    ("other", "other (as specified"),
    ("international_bunkers.total", "international bunkers"),
    ("international_bunkers.aviation", "aviation"),
    ("international_bunkers.navigation", "navigation"),
    ("multilateral_operations", "multilateral operations"),
    ("co2_emissions_from_biomass", "co2 emissions from biomass"),
    ("co2_captured", "co2 captured"),
    ("longerim_storage_waste_disposal", "long-term storage of c in waste disposal"),
    ("indirect_n20", "indirect n2o"),
    ("indirect_co2", "indirect co2"),
    ("total_source", "total co2 equivalent emissions without"),
]

TRANSPORT_SEQUENCE = [
    ("total", "transport"),
    ("domestic_aviation", "domestic aviation"),
    ("road_transporation.total", "road transportation"),
    ("road_transporation.cars", "cars"),
    ("road_transporation.light_duty_trucks", "light duty trucks"),
    # the workbook only ever reports this as one combined row (see module
    # docstring) - the trucks/buses split stored separately below comes from
    # a different, manually-maintained source.
    ("road_transporation.heavy_duty_trucks_and_buses", "heavy duty trucks and buses"),
    ("road_transporation.motorcycles", "motorcycles"),
    ("road_transporation.other", "other"),
    ("railways", "railways"),
    ("domestic_navigation", "domestic navigation"),
    ("other_transportation", "other transportation"),
]


def extract_summary2(path):
    wb = xl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Summary2"]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
        # labels in columns A-J (0-9), value in column K (10)
        return match_sequence(rows, SUMMARY2_SEQUENCE, range(10), lambda row: row[10].value, path)
    finally:
        wb.close()


def extract_transport(path, year, summary2_values, truck_bus_split):
    wb = xl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Table1.A(a)s3"]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
        # label in column B (1); CO2/CH4/N2O emissions (kt) in columns H/I/J (7/8/9)
        values = match_sequence(
            rows, TRANSPORT_SEQUENCE, [1],
            lambda row: co2equiv(row[7].value, row[8].value, row[9].value),
            path,
        )
    finally:
        wb.close()

    heavy_duty_trucks, buses = truck_bus_split.get(year, (0.0, 0.0))
    values["road_transporation.heavy_duty_trucks"] = heavy_duty_trucks
    values["road_transporation.buses"] = buses
    values["international_aviation"] = to_number(summary2_values["international_bunkers.aviation"])
    values["international_navigation"] = to_number(summary2_values["international_bunkers.navigation"])
    return values


TRUCK_BUS_SPLIT_FILE_GLOB = "../sources/Emisije TGP iz cestnega prometa*.xlsx"
TRUCK_BUS_SPLIT_YEAR_START = 1986
TRUCK_BUS_SPLIT_YEAR_END = 2021


def load_truck_bus_split():
    """Reads the manually-maintained heavy-duty-truck/bus CO2-eq split
    (rows 4-9: alternating heavy-trucks/buses rows for CO2, CH4, N2O; columns
    starting at 1986 in column 4). Covers TRUCK_BUS_SPLIT_YEAR_START through
    TRUCK_BUS_SPLIT_YEAR_END only - see module docstring."""
    matches = glob.glob(TRUCK_BUS_SPLIT_FILE_GLOB)
    if not matches:
        print(f"WARNING: truck/bus split file not found ({TRUCK_BUS_SPLIT_FILE_GLOB}); "
              f"heavy_duty_trucks/buses will be 0.0 for all years")
        return {}
    wb = xl.load_workbook(matches[0], data_only=True, read_only=True)
    try:
        ws = wb.active
        split = {}
        for year in range(TRUCK_BUS_SPLIT_YEAR_START, TRUCK_BUS_SPLIT_YEAR_END + 1):
            col = year - 1982
            heavy = co2equiv(ws.cell(4, col).value, ws.cell(6, col).value, ws.cell(8, col).value)
            buses = co2equiv(ws.cell(5, col).value, ws.cell(7, col).value, ws.cell(9, col).value)
            split[year] = (heavy, buses)
        return split
    finally:
        wb.close()


def backup_current_data(data_dir):
    timestamp = datetime.today().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(data_dir, f"backup_{timestamp}")
    os.makedirs(backup_dir, exist_ok=True)
    for csv_path in glob.glob(os.path.join(data_dir, "*.csv")):
        shutil.copy2(csv_path, backup_dir)
    print(f"Backed up current CSVs to {backup_dir}")
    return backup_dir


def main():
    sources_dir = "../sources/SVN_2026"
    data_dir = "../data"

    backup_current_data(data_dir)

    year_files = {}
    for path in glob.glob(os.path.join(sources_dir, "SVN-CRT-2026-V1.0-*.xlsx")):
        m = re.search(r"SVN-CRT-2026-V1\.0-(\d{4})\.xlsx$", os.path.basename(path))
        if not m:
            continue
        year = int(m.group(1))
        if YEAR_START <= year <= YEAR_END:
            year_files[year] = path

    missing = sorted(set(range(YEAR_START, YEAR_END + 1)) - set(year_files))
    if missing:
        raise RuntimeError(f"Missing SVN_2026 source file(s) for year(s): {missing}")

    years = sorted(year_files)
    print(f"Processing {len(years)} years: {years[0]}-{years[-1]}")

    truck_bus_split = load_truck_bus_split()

    summary2_by_year = {}
    transport_by_year = {}
    for year in years:
        path = year_files[year]
        print(year, path)
        summary2_by_year[year] = extract_summary2(path)
        transport_by_year[year] = extract_transport(path, year, summary2_by_year[year], truck_bus_split)

    def series(extractor):
        return np.array([extractor(summary2_by_year[y]) for y in years])

    def s2(key):
        return series(lambda v: to_number(v[key]))

    years_arr = np.array(years)

    df = pd.DataFrame({
        "year": years_arr,
        "total_source": s2("total_source"),
        "energy_industries": s2("energy.fuel_combustion_activities.energy_industries"),
        "manufacturing_construction_fuels": s2("energy.fuel_combustion_activities.manufacturing_construction"),
        "transport": s2("energy.fuel_combustion_activities.transport"),
        "industrial_processes": s2("industrial_processes.total"),
        "residential_commercial_agricultural_forestry_fishing_fuels": s2("energy.fuel_combustion_activities.other_sectors"),
        "agriculture": s2("agriculture.total"),
        "waste": s2("waste.total"),
        "international_aviation": s2("international_bunkers.aviation"),
        "international_navigation": s2("international_bunkers.navigation"),
        "co2_emissions_from_biomass": s2("co2_emissions_from_biomass"),
        "others": s2("energy.fuel_combustion_activities.other") + s2("energy.fugitive_emissions_from_fuels.total"),
        "lulucf": s2("lulucf.total"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.csv"), index=False, float_format="%.2f")

    df = pd.DataFrame({
        "year": years_arr,
        "total": s2("energy.total"),
        "fuel_combustion_activities.total": s2("energy.fuel_combustion_activities.total"),
        "fuel_combustion_activities.energy_industries": s2("energy.fuel_combustion_activities.energy_industries"),
        "fuel_combustion_activities.manufacturing_construction": s2("energy.fuel_combustion_activities.manufacturing_construction"),
        "fuel_combustion_activities.transport": s2("energy.fuel_combustion_activities.transport"),
        "fuel_combustion_activities.other_sectors": s2("energy.fuel_combustion_activities.other_sectors"),
        "fuel_combustion_activities.other": s2("energy.fuel_combustion_activities.other"),
        "fugitive_emissions_from_fuels.total": s2("energy.fugitive_emissions_from_fuels.total"),
        "fugitive_emissions_from_fuels.solid_fuels": s2("energy.fugitive_emissions_from_fuels.solid_fuels"),
        "fugitive_emissions_from_fuels.oil_natural_gas_and_energy_production": s2("energy.fugitive_emissions_from_fuels.oil_natural_gas_and_energy_production"),
        "co2_transport_storage": s2("energy.co2_transport_storage"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.energy.csv"), index=False, float_format="%.2f")

    df = pd.DataFrame({
        "year": years_arr,
        "total": s2("industrial_processes.total"),
        "mineral_industry": s2("industrial_processes.mineral_industry"),
        "chemical_industry": s2("industrial_processes.chemical_industry"),
        "metal_industry": s2("industrial_processes.metal_industry"),
        "non_energy_products_from_fuels": s2("industrial_processes.non_energy_products_from_fuels"),
        "electronic_industry": s2("industrial_processes.electronic_industry"),
        "product_usese_as_ODS": s2("industrial_processes.product_usese_as_ODS"),
        "other_product_manufacture_use": s2("industrial_processes.other_product_manufacture_use"),
        "other": s2("industrial_processes.other"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.industrial_processes.csv"), index=False, float_format="%.2f")

    df = pd.DataFrame({
        "year": years_arr,
        "total": s2("agriculture.total"),
        "enteric_fermentation": s2("agriculture.enteric_fermentation"),
        "manure_management": s2("agriculture.manure_management"),
        "rice_cultivation": s2("agriculture.rice_cultivation"),
        "agricultural_soils": s2("agriculture.agricultural_soils"),
        "prescribed_burning_of_savannas": s2("agriculture.prescribed_burning_of_savannas"),
        "field_burning_agricultural_residues": s2("agriculture.field_burning_agricultural_residues"),
        "liming": s2("agriculture.liming"),
        "urea_application": s2("agriculture.urea_application"),
        "carbon_containing_fertilizers": s2("agriculture.carbon_containing_fertilizers"),
        "other": s2("agriculture.other"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.agriculture.csv"), index=False, float_format="%.2f")

    df = pd.DataFrame({
        "year": years_arr,
        "total": s2("lulucf.total"),
        "forest_land": s2("lulucf.forest_land"),
        "cropland": s2("lulucf.cropland"),
        "grassland": s2("lulucf.grassland"),
        "wetlands": s2("lulucf.wetlands"),
        "settlements": s2("lulucf.settlements"),
        "other_land": s2("lulucf.other_land"),
        "harvested_wood_prducts": s2("lulucf.harvested_wood_prducts"),
        "other": s2("lulucf.other"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.lulucf.csv"), index=False, float_format="%.2f")

    df = pd.DataFrame({
        "year": years_arr,
        "total": s2("waste.total"),
        "solid_waste_disposal": s2("waste.solid_waste_disposal"),
        "biological_treatment_solid_waste": s2("waste.biological_treatment_solid_waste"),
        "incineration_open_burning_waste": s2("waste.incineration_open_burning_waste"),
        "waste_water_treatment_discharge": s2("waste.waste_water_treatment_discharge"),
        "other": s2("waste.other"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.waste.csv"), index=False, float_format="%.2f")

    df = pd.DataFrame({
        "year": years_arr,
        "international_bunkers.total": s2("international_bunkers.total"),
        "international_bunkers.aviation": s2("international_bunkers.aviation"),
        "international_bunkers.navigation": s2("international_bunkers.navigation"),
        "multilateral_operations": s2("multilateral_operations"),
        "co2_emissions_from_biomass": s2("co2_emissions_from_biomass"),
        "co2_captured": s2("co2_captured"),
        "longerim_storage_waste_disposal": s2("longerim_storage_waste_disposal"),
        "indirect_n20": s2("indirect_n20"),
        "indirect_co2": s2("indirect_co2"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.memo_items.csv"), index=False, float_format="%.2f")

    def tr(key):
        return np.array([transport_by_year[y][key] for y in years])

    df = pd.DataFrame({
        "year": years_arr,
        "total": tr("total"),
        "road_transporation.total": tr("road_transporation.total"),
        "road_transporation.cars": tr("road_transporation.cars"),
        "road_transporation.light_duty_trucks": tr("road_transporation.light_duty_trucks"),
        "road_transporation.heavy_duty_trucks": tr("road_transporation.heavy_duty_trucks"),
        "road_transporation.buses": tr("road_transporation.buses"),
        "road_transporation.heavy_duty_trucks_and_buses": tr("road_transporation.heavy_duty_trucks_and_buses"),
        "road_transporation.motorcycles": tr("road_transporation.motorcycles"),
        "road_transporation.other": tr("road_transporation.other"),
        "railways": tr("railways"),
        "domestic_aviation": tr("domestic_aviation"),
        "domestic_navigation": tr("domestic_navigation"),
        "other_transportation": tr("other_transportation"),
        "international_aviation": tr("international_aviation"),
        "international_navigation": tr("international_navigation"),
    })
    df.to_csv(os.path.join(data_dir, "emissions.historical.energy.transport.csv"), index=False, float_format="%.2f")

    print(f"Wrote 8 CSVs to {data_dir} for years {years[0]}-{years[-1]}.")
    no_split_years = [y for y in years if y > TRUCK_BUS_SPLIT_YEAR_END]
    if no_split_years:
        print(
            f"Note: road_transporation.heavy_duty_trucks/.buses are 0.0 for "
            f"{no_split_years[0]}-{no_split_years[-1]} - the manually-maintained "
            f"split spreadsheet doesn't cover those years; update it separately "
            f"and re-run if needed. road_transporation.heavy_duty_trucks_and_buses "
            f"(the combined figure straight from the CRT workbooks) is populated "
            f"for those years regardless."
        )


if __name__ == "__main__":
    main()
